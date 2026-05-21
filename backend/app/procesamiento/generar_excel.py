import os
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import io
from typing import Dict, List, Optional, Tuple
from datetime import datetime

ESTILO_CONCILIADO = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
ESTILO_NO_CONCILIADO = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
ESTILO_CABECERA = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
ESTILO_BLANCO = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
ESTILO_DIFERENCIA = PatternFill(start_color="9CB2D4", end_color="9CB2D4", fill_type="solid")

FUENTE_NEGRITA = Font(bold=True)
FUENTE_ROJA = Font(color="FF0000")
FUENTE_VERDE = Font(color="2ECC71")
FUENTE_NEGRITA_ROJA = Font(bold=True, color="FF0000")
FUENTE_NORMAL = Font(bold=False)

ALINEACION_ESTANDAR = Alignment(wrap_text=True, vertical='top', horizontal='left')
ALINEACION_CENTRO = Alignment(horizontal='center', vertical='center', wrap_text=True)
ALINEACION_DERECHA = Alignment(horizontal='right', vertical='center', wrap_text=True)

BORDE_GRUESO = Border(
    left=Side(style='medium', color="000000"),
    right=Side(style='medium', color="000000"),
    top=Side(style='medium', color="000000"),
    bottom=Side(style='medium', color="000000")
)

def crear_excel_actualizado(
    contenido_excel: bytes,
    nombre_hoja: str,
    movimientos_nuevos: List[Dict],
    resultado_conciliacion: Dict,
    mes: int,
    año: int,
    nombre_documento: str = "",
    es_excel: bool = True
) -> bytes:
    # 1. Determinar nombre de hoja dinámico basado en las transacciones
    # IMPORTANTE: no reasignar nombre_hoja en modo histórico si venía ya calculado.
    # Si el string FECHA no corresponde al mes/año esperado, puedes acabar creando una hoja "incorrecta"
    # y el usuario percibe que el histórico "no se copia".
    if movimientos_nuevos and not nombre_hoja:
        for mov in movimientos_nuevos:
            fecha_str = mov.get("FECHA")
            if fecha_str and len(str(fecha_str)) >= 10:
                try:
                    dt = datetime.strptime(str(fecha_str), "%d/%m/%Y")
                    meses_esp = ["ENERO", "FEBRERO", "MARZO", "ABRIL", "MAYO", "JUNIO",
                                 "JULIO", "AGOSTO", "SEPTIEMBRE", "OCTUBRE", "NOVIEMBRE", "DICIEMBRE"]
                    nombre_hoja = f"{meses_esp[dt.month-1]} {dt.year}"
                    break
                except:
                    continue

    if not nombre_hoja:
        raise ValueError("nombre_hoja no puede estar vacío")

    # Definir encabezados dinámicamente según la presencia de ORDENANTE
    headers = ["FECHA", "OBSERVACIONES", "IMPORTE", "SALDO", "CONCEPTO"]
    if es_excel: # Si el archivo original era Excel, siempre incluimos ORDENANTE
        headers.insert(1, "ORDENANTE")
    
    num_cols = len(headers)

    # Si no se provee contenido, creamos un libro nuevo (Modo solo extracto)
    if not contenido_excel:
        workbook = openpyxl.Workbook()
        hoja = workbook.active
        hoja.title = nombre_hoja
    else:
        try:
            workbook = load_workbook(io.BytesIO(contenido_excel))
        except Exception as e:
            raise ValueError(f"Error al cargar Excel histórico: {str(e)}")
        
        # Determinar el nombre final de la hoja para evitar duplicados
        final_nombre_hoja = nombre_hoja
        if final_nombre_hoja in workbook.sheetnames:
            # Si la hoja ya existe, creamos una nueva con un sufijo para preservar la original
            counter = 1
            while f"{nombre_hoja} ({counter})" in workbook.sheetnames:
                counter += 1
            final_nombre_hoja = f"{nombre_hoja} ({counter})"
        
        # Siempre la creamos al final para que sea "la última" y contenga todas las originales
        hoja = workbook.create_sheet(final_nombre_hoja)

    # Establecer la hoja recién creada como activa para que el Excel se abra por ella
    workbook.active = workbook.index(hoja)

    # --- Títulos y Encabezados (Header 1 y Header 2) ---
    # Combinamos celdas a lo ancho de la tabla para que el nombre sea visible
    # Usamos el nombre directo sin splitext para evitar que nombres de comunidades con puntos se corten
    nombre_entidad = str(nombre_documento).upper()
    
    # Header 1: Nombre del Documento / Comunidad
    hoja.merge_cells(start_row=1, start_column=1, end_row=1, end_column=num_cols)
    cell_doc = hoja.cell(row=1, column=1, value=nombre_entidad)
    cell_doc.font = Font(bold=True, size=16)
    cell_doc.alignment = ALINEACION_CENTRO
    hoja.row_dimensions[1].height = 25

    # Header 2: Nombre de la Hoja (Mes y Año) o Nombre de Comunidad si se prefiere aquí
    hoja.merge_cells(start_row=2, start_column=1, end_row=2, end_column=num_cols)
    cell_sheet = hoja.cell(row=2, column=1, value=nombre_hoja)
    cell_sheet.font = Font(bold=True, size=14, color="34495E")
    cell_sheet.alignment = ALINEACION_CENTRO
    hoja.row_dimensions[2].height = 22

    for col_idx, text in enumerate(headers, 1):
        cell = hoja.cell(row=4, column=col_idx, value=text)
        cell.font = FUENTE_NEGRITA
        cell.alignment = ALINEACION_CENTRO
        cell.fill = ESTILO_CABECERA
        cell.border = BORDE_GRUESO

    # La cabecera termina en la fila 4, los datos empiezan en la 5
    ultima_fila = 4
    
    # Usar movimientos_nuevos directamente (los enviados desde la UI editada)
    for mov_extracto in movimientos_nuevos:
        if not mov_extracto:
            continue
        
        ultima_fila += 1
        
        is_gasto = float(mov_extracto.get("IMPORTE", 0)) < 0

        for col_idx, header in enumerate(headers, 1):
            val = mov_extracto.get(header, "")
            if header == "CONCEPTO" and (not val or str(val).lower() == "nan"):
                val = "Sin asignar"
            
            cell = hoja.cell(row=ultima_fila, column=col_idx, value=val)
            cell.fill = ESTILO_BLANCO
            cell.border = BORDE_GRUESO
            
            # Aplicar alineación según requerimiento:
            # Excel: ORDENANTE y OBSERVACIONES a la izquierda, resto centradas.
            # CSV: OBSERVACIONES a la izquierda, resto centradas.
            if header == "OBSERVACIONES" or (es_excel and header == "ORDENANTE"):
                cell.alignment = ALINEACION_ESTANDAR
            else:
                cell.alignment = ALINEACION_CENTRO
            
            # Estilos de fuente (Rojo para gastos, Negrita para concepto)
            is_col_concepto = (header == "CONCEPTO")
            if is_gasto:
                cell.font = FUENTE_NEGRITA_ROJA if is_col_concepto else FUENTE_ROJA
            elif is_col_concepto:
                cell.font = FUENTE_NEGRITA
            else:
                cell.font = FUENTE_NORMAL

    # Ajustar anchos de columna automáticamente
    for col in hoja.columns:
        max_length = 0
        # Usamos get_column_letter con el índice numérico para evitar errores con MergedCells
        column_letter = get_column_letter(col[0].column)
        for cell in col:
            # Solo calculamos el ancho basado en los datos de la tabla (fila 4 en adelante)
            if cell.row >= 4 and cell.value:
                max_length = max(max_length, len(str(cell.value)))
        
        # Establecemos un ancho mínimo de seguridad y un máximo para que no sea infinito
        hoja.column_dimensions[column_letter].width = min(max_length + 4, 60)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    
    return output.getvalue()

def crear_excel_informe_finanzas(
    nombre_documento: str,
    nombre_hoja: str,
    finanzas_data: Dict
) -> bytes:
    workbook = openpyxl.Workbook()
    hoja = workbook.active
    hoja.title = nombre_hoja

    # Definición de columnas por tabla
    num_cols_ingresos = 3  # Piso, Importe, Fecha
    num_cols_gastos = 2    # Concepto, Cantidad
    num_cols_resumen = 2   # Concepto, Importe


    # --- Títulos Generales ---
    nombre_entidad = str(nombre_documento).upper()
    
    hoja.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(num_cols_ingresos, num_cols_gastos, num_cols_resumen))
    cell_doc = hoja.cell(row=1, column=1, value=nombre_entidad)
    cell_doc.font = Font(bold=True, size=16)
    cell_doc.alignment = ALINEACION_CENTRO
    hoja.row_dimensions[1].height = 25

    hoja.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(num_cols_ingresos, num_cols_gastos, num_cols_resumen))
    cell_sheet = hoja.cell(row=2, column=1, value=nombre_hoja)
    cell_sheet.font = Font(bold=True, size=14, color="34495E")
    cell_sheet.alignment = ALINEACION_CENTRO
    hoja.row_dimensions[2].height = 22

    current_row = 4
    # --- Tabla 1: Ingresos por pisos ---
    hoja.cell(row=current_row, column=1, value="INGRESOS POR PISOS").font = Font(bold=True, size=12, color="000000")
    current_row += 1

    headers_ingresos = ["Piso", "Fecha", "Importe"]

    for col_idx, text in enumerate(headers_ingresos, 1):
        cell = hoja.cell(row=current_row, column=col_idx, value=text)
        cell.font = FUENTE_NEGRITA
        cell.alignment = ALINEACION_CENTRO
        cell.fill = ESTILO_CABECERA
        cell.border = BORDE_GRUESO
    current_row += 1

    # Totales generales de la tabla (no por fila)
    ingresos_total_mes = 0.0

    # Sumar importes para calcular el total de ingresos del mes
    for item in finanzas_data.get("ingresosPorPiso", []):
        imp = item.get("importe")
        if imp is None or (isinstance(imp, str) and imp.strip() == ""):
            continue
        try:
            ingresos_total_mes += float(imp)
        except:
            continue

    for item in finanzas_data.get("ingresosPorPiso", []):
        # Poner bordes también en columnas/filas de datos
        hoja.cell(row=current_row, column=1, value=item.get("codigo")).alignment = ALINEACION_ESTANDAR

        hoja.cell(row=current_row, column=2, value=item.get("fecha")).alignment = ALINEACION_CENTRO

        importe_val = item.get("importe")
        if importe_val is None or (isinstance(importe_val, str) and importe_val.strip() == ""):
            cell_vacio = hoja.cell(row=current_row, column=3, value=" ")
            cell_vacio.alignment = ALINEACION_CENTRO
        else:
            cell_imp = hoja.cell(row=current_row, column=3, value=importe_val)
            cell_imp.number_format = '#,##0.00"€"'
            cell_imp.alignment = ALINEACION_CENTRO

        # Bordes de la fila de datos
        for col_idx in range(1, 4):
            c = hoja.cell(row=current_row, column=col_idx)
            c.border = BORDE_GRUESO

        current_row += 1

    # Fila de Total Ingresos
    hoja.cell(row=current_row, column=1, value="TOTAL INGRESOS").font = FUENTE_NEGRITA
    hoja.cell(row=current_row, column=3, value=round(ingresos_total_mes, 2)).number_format = '#,##0.00"€"'
    hoja.cell(row=current_row, column=3).alignment = ALINEACION_CENTRO
    hoja.cell(row=current_row, column=3).font = FUENTE_NEGRITA
    for col_idx in range(1, 4):
        c = hoja.cell(row=current_row, column=col_idx)
        c.border = BORDE_GRUESO
    current_row += 1


    current_row += 2 # Espacio entre tablas


    # --- Tabla 2: Gastos del Mes ---
    hoja.cell(row=current_row, column=1, value="GASTOS DEL MES").font = Font(bold=True, size=12, color="E74C3C")
    current_row += 1

    headers_gastos = ["Concepto", "Cantidad"]

    for col_idx, text in enumerate(headers_gastos, 1):
        cell = hoja.cell(row=current_row, column=col_idx, value=text)
        cell.font = FUENTE_NEGRITA
        cell.alignment = ALINEACION_CENTRO
        cell.fill = ESTILO_CABECERA
        cell.border = BORDE_GRUESO
    current_row += 1

    gastos_total_mes = 0.0
    for item in finanzas_data.get("gastos", []):
        hoja.cell(row=current_row, column=1, value=item.get("concepto")).alignment = ALINEACION_ESTANDAR
        hoja.cell(row=current_row, column=2, value=item.get("importe")).number_format = '#,##0.00"€"'
        imp_gasto = item.get("importe", 0)
        hoja.cell(row=current_row, column=2, value=imp_gasto).number_format = '#,##0.00"€"'
        hoja.cell(row=current_row, column=2).alignment = ALINEACION_DERECHA
        gastos_total_mes += float(imp_gasto)
        for col_idx in range(1, 3):
            hoja.cell(row=current_row, column=col_idx).border = BORDE_GRUESO
        current_row += 1

    # Fila de Total Gastos
    hoja.cell(row=current_row, column=1, value="TOTAL GASTOS").font = FUENTE_NEGRITA
    hoja.cell(row=current_row, column=2, value=round(gastos_total_mes, 2)).number_format = '#,##0.00"€"'
    hoja.cell(row=current_row, column=2).alignment = ALINEACION_DERECHA
    hoja.cell(row=current_row, column=2).font = FUENTE_NEGRITA
    for col_idx in range(1, 3):
        hoja.cell(row=current_row, column=col_idx).border = BORDE_GRUESO
    current_row += 1
    
    current_row += 2 # Espacio entre tablas

    # --- Tabla 3: Resumen de Cuentas (Tabla final de cálculo) ---
    hoja.cell(row=current_row, column=1, value="RESUMEN DE CUENTAS").font = Font(bold=True, size=12, color="6366F1")
    current_row += 1

    resumen_cuentas = finanzas_data.get("resumenCuentas", {})

    headers_resumen = ["Concepto", "Importe (€)"]

    for col_idx, text in enumerate(headers_resumen, 1):
        cell = hoja.cell(row=current_row, column=col_idx, value=text)
        cell.font = FUENTE_NEGRITA
        cell.alignment = ALINEACION_CENTRO
        cell.fill = ESTILO_CABECERA
        cell.border = BORDE_GRUESO
    current_row += 1

    # Filas (estructura pedida para el informe)
    # - Debajo de "Gastos" mostramos el total calculado del periodo:
    #   saldo_mes = saldoAnterior + ingresosMes - gastosMes
    saldo_anterior = resumen_cuentas.get("saldoAnterior", 0)
    ingresos_mes = resumen_cuentas.get("ingresosMes", 0)
    gastos_mes = resumen_cuentas.get("gastosMes", 0)
    saldo_mes = saldo_anterior + ingresos_mes - gastos_mes

    filas_resumen = [
        ("Saldo Anterior", saldo_anterior, "Saldo antes de este mes"),
        ("Ingresos", ingresos_mes, "Suma de ingresos del mes"),
        ("Gastos", gastos_mes, "Suma de gastos del mes"),
        ("Total saldo mes", saldo_mes, "Saldo final del mes")
    ]

    for concepto, importe, _significado in filas_resumen:
        cell_concepto = hoja.cell(row=current_row, column=1, value=concepto)
        cell_importe = hoja.cell(row=current_row, column=2, value=importe)
        
        cell_concepto.alignment = ALINEACION_ESTANDAR
        cell_importe.alignment = ALINEACION_DERECHA
        
        # Formato numérico base
        cell_importe.number_format = '#,##0.00"€"'

        if concepto == "Ingresos":
            cell_importe.font = FUENTE_VERDE
            cell_importe.number_format = '"+"#,##0.00"€"'
        elif concepto == "Gastos":
            cell_importe.font = FUENTE_ROJA
            cell_importe.number_format = '"-"#,##0.00"€"'
        elif concepto == "Total saldo mes":
            cell_concepto.font = FUENTE_NEGRITA
            cell_importe.font = FUENTE_NEGRITA

        # Bordes (solo 2 columnas)
        for col_idx in range(1, 3):
            c = hoja.cell(row=current_row, column=col_idx)
            c.border = BORDE_GRUESO

        current_row += 1



    # Ajustar anchos de columna automáticamente para las columnas usadas
    for col_idx in range(1, max(num_cols_ingresos, num_cols_gastos, num_cols_resumen) + 1):
        max_length = 0
        for row_idx in range(1, hoja.max_row + 1):
            cell_value = hoja.cell(row=row_idx, column=col_idx).value
            if cell_value:
                max_length = max(max_length, len(str(cell_value)))
        hoja.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 4, 60)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    
    return output.getvalue()

def recalcular_totales(hoja: openpyxl.worksheet.worksheet.Worksheet) -> openpyxl.worksheet.worksheet.Worksheet:
    ultima_fila = hoja.max_row
    
    fila_totales = None
    for row in range(1, ultima_fila + 1):
        valor = hoja.cell(row=row, column=1).value
        if valor and "TOTAL" in str(valor).upper():
            fila_totales = row
            break
    
    if not fila_totales:
        fila_totales = ultima_fila + 2
        hoja.cell(row=fila_totales, column=1, value="TOTALES")
        hoja.cell(row=fila_totales, column=1).font = FUENTE_NEGRITA
    
    suma_importe = 0
    
    for row in range(2, fila_totales):
        imp_val = hoja.cell(row=row, column=3).value
        
        try:
            if imp_val:
                suma_importe += float(imp_val)
        except:
            pass
        
    # Mostrar el total de movimientos en la columna de importe
    hoja.cell(row=fila_totales, column=3, value=round(suma_importe, 2))
    hoja.cell(row=fila_totales, column=3).font = FUENTE_NEGRITA
    
    # El saldo final del periodo en la columna 4
    hoja.cell(row=fila_totales, column=4).font = FUENTE_NEGRITA
    
    return hoja

def crear_excel_resumen(
    mes: int,
    año: int,
    resultado_conciliacion: Dict
) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    
    wb = Workbook()
    ws = wb.active
    ws.title = "Resumen Conciliación"
    
    meses = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
             "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    
    ws.cell(row=1, column=1, value=f"Resumen de Conciliación - {meses[mes-1]} {año}")
    ws.cell(row=1, column=1).font = Font(bold=True, size=14)
    
    resumen = resultado_conciliacion.get("resumen", {})
    
    row = 3
    ws.cell(row=row, column=1, value="CONCILIADOS")
    ws.cell(row=row, column=1).font = Font(bold=True)
    
    row += 1
    ws.cell(row=row, column=1, value="Movimientos conciliados:")
    ws.cell(row=row, column=2, value=resumen.get("conciliados", 0))
    
    row += 1
    ws.cell(row=row, column=1, value="Ingresos conciliados:")
    ws.cell(row=row, column=2, value=resumen.get("ingresos_conciliados", 0))
    
    row += 1
    ws.cell(row=row, column=1, value="Gastos conciliados:")
    ws.cell(row=row, column=2, value=resumen.get("gastos_conciliados", 0))
    
    row += 2
    ws.cell(row=row, column=1, value="NO CONCILIADOS")
    ws.cell(row=row, column=1).font = Font(bold=True)
    
    row += 1
    ws.cell(row=row, column=1, value="Movimientos nuevos:")
    ws.cell(row=row, column=2, value=resumen.get("no_conciliados", 0))
    
    row += 1
    ws.cell(row=row, column=1, value="Ingresos nuevos:")
    ws.cell(row=row, column=2, value=resumen.get("ingresos_nuevos", 0))
    
    row += 1
    ws.cell(row=row, column=1, value="Gastos nuevos:")
    ws.cell(row=row, column=2, value=resumen.get("gastos_nuevos", 0))
    
    row += 2
    ws.cell(row=row, column=1, value="DIFERENCIAS")
    ws.cell(row=row, column=1).font = Font(bold=True)
    
    row += 1
    ws.cell(row=row, column=1, value="Diferencias encontradas:")
    ws.cell(row=row, column=2, value=resumen.get("diferencias", 0))
    
    ws.column_dimensions['A'].width = 30
    ws.column_dimensions['B'].width = 20
    
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()

def generar_excel_descarga(
    contenido_excel: bytes,
    nombre_hoja: str,
    resultado_conciliacion: Dict,
    movimientos_extracto: List[Dict],
    mes: int,
    año: int
) -> Tuple[bytes, str]:
    excel_bytes = crear_excel_actualizado(
        contenido_excel,
        nombre_hoja,
        [],
        resultado_conciliacion,
        mes,
        año
    )
    
    meses = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
            "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    
    nombre_archivo = f"Contabilidad_{meses[mes-1]}_{año}_Actualizado.xlsx"
    
    return excel_bytes, nombre_archivo

GenerarExcelActualizado = crear_excel_actualizado
GenerarExcelResumen = crear_excel_resumen