"""
procesa el Excel contable
Lee el Excel con hojas por cada mes y extrae ingresos y gastos.
"""
import pandas as pd
import openpyxl
from openpyxl import load_workbook
from typing import Dict, List, Optional, Tuple
import re
from datetime import datetime


def obtener_nombre_hoja(mes: int, año: int) -> str:
    """
    Genera el nombre de la hoja según el mes y año.
    Formatos soportados: "Enero 2024", "2024-01", "01_2024", etc.
    """
    meses = [
        "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
        "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
    ]
    
    nombres = [
        f"{meses[mes-1]} {año}",
        f"{año}-{mes:02d}",
        f"{mes:02d}_{año}",
        f"{meses[mes-1][:3].upper()} {año}",
        str(mes),
    ]
    
    return nombres[0]


def detectar_hoja_por_mes(workbook: openpyxl.Workbook, mes: int, año: int) -> Optional[str]:
    meses = [
        "Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
        "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"
    ]
    
    mes_nombre = meses[mes - 1]
    mes_corto = mes_nombre[:3]
    
    for hoja in workbook.sheetnames:
        hoja_lower = hoja.lower()

        if str(mes) in hoja_lower and str(año) in hoja_lower:
            return hoja
        if mes_nombre.lower() in hoja_lower and str(año) in hoja_lower:
            return hoja
        if mes_corto.lower() in hoja_lower and str(año) in hoja_lower:
            return hoja
        # Solo coincide con año
        if str(año) in hoja_lower:
            return hoja
    
    return workbook.sheetnames[0] if workbook.sheetnames else None


def detectar_columnas_contables(df: pd.DataFrame) -> Dict[str, str]:
    """
    Detecta las columnas del Excel contable.
    """
    cols = list(df.columns)
    resultado = {
        "fecha": None,
        "concepto": None,
        "importe": None,
        "tipo": None,
        "categoria": None,
        "debe": None,
        "haber": None,
    }
    
    for col in cols:
        col_lower = col.lower().strip()
        
        if resultado["fecha"] is None and ("fecha" in col_lower or "date" in col_lower):
            resultado["fecha"] = col
        
        if resultado["concepto"] is None and (
            "concepto" in col_lower
            or "descripcion" in col_lower
            or "observaciones" in col_lower
            or "detalle" in col_lower
        ):
            resultado["concepto"] = col
        
        if resultado["importe"] is None and ("importe" in col_lower or "amount" in col_lower):
            resultado["importe"] = col
        
        if resultado["tipo"] is None and ("tipo" in col_lower or "type" in col_lower):
            resultado["tipo"] = col
        
        if resultado["categoria"] is None and ("categoria" in col_lower or "category" in col_lower):
            resultado["categoria"] = col
        
        if resultado["debe"] is None and ("debe" in col_lower or "cargo" in col_lower or "dr" in col_lower):
            resultado["debe"] = col
        if resultado["haber"] is None and ("haber" in col_lower or "abono" in col_lower or "cr" in col_lower):
            resultado["haber"] = col
    
    if resultado["debe"] and resultado["haber"]:
        resultado["importe"] = None  # Se calculan dinámicamente
    
    return resultado


def limpiar_importe(valor) -> float:
    """Limpia y convierte un importe a float."""
    if pd.isna(valor):
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)
    
    texto = str(valor).strip()
    texto = texto.replace(".", "").replace(",", ".")
    texto = re.sub(r"[^\d.\-]", "", texto)
    
    try:
        return float(texto)
    except:
        return 0.0


def determinar_tipo(importe: float, row: pd.Series, columnas: Dict[str, str]) -> str:
    """
    Determina el tipo de movimiento (ingreso/gasto).
    """
    if columnas.get("tipo") and columnas["tipo"] in row.index:
        tipo_val = str(row[columnas["tipo"]]).lower()
        if "ingreso" in tipo_val or "haber" in tipo_val or "abono" in tipo_val:
            return "ingreso"
        if "gasto" in tipo_val or "debe" in tipo_val or "cargo" in tipo_val:
            return "gasto"
    
    if columnas.get("debe") and columnas.get("haber"):
        debe = limpiar_importe(row.get(columnas["debe"], 0))
        haber = limpiar_importe(row.get(columnas["haber"], 0))
        
        if haber > 0:
            return "ingreso"
        if debe > 0:
            return "gasto"
    
    if importe > 0:
        return "ingreso"
    else:
        return "gasto"


def leer_excel_contable(
    contenido: bytes,
    mes: Optional[int] = None,
    año: Optional[int] = None
) -> Tuple[List[Dict], Dict]:
    
    try:
        workbook = load_workbook(io.BytesIO(contenido), data_only=True)
    except:
        df = pd.read_excel(io.BytesIO(contenido), sheet_name=None)
        first_sheet = list(df.keys())[0]
        return [], {"error": "No se pudo leer el Excel"}
    
    if mes and año:
        nombre_hoja = detectar_hoja_por_mes(workbook, mes, año)
    else:
        nombre_hoja = workbook.sheetnames[0] if workbook.sheetnames else None
    
    if not nombre_hoja:
        return [], {"error": "No se encontró hoja"}
    
    try:
        df = pd.read_excel(io.BytesIO(contenido), sheet_name=nombre_hoja)
    except Exception as e:
        return [], {"error": f"Error al leer hoja: {str(e)}"}
    
    columnas = detectar_columnas_contables(df)
    
    if columnas["concepto"] is None:
        return [], {"error": "No se encontró columna de concepto"}
    
    movimientos = []
    total_ingresos = 0.0
    total_gastos = 0.0
    num_ingresos = 0
    num_gastos = 0
    
    for idx, row in df.iterrows():
        concepto = str(row.get(columnas["concepto"], "")).strip()
        
        if not concepto or concepto == "nan":
            continue
        
        if columnas.get("debe") and columnas.get("haber"):
            debe = limpiar_importe(row.get(columnas["debe"], 0))
            haber = limpiar_importe(row.get(columnas["haber"], 0))
            importe = haber - debe
        else:
            importe = limpiar_importe(row.get(columnas["importe"], 0))
        
        if importe == 0:
            continue
        
        tipo = determinar_tipo(importe, row, columnas)
        
        fecha = None
        if columnas.get("fecha") and columnas["fecha"] in row.index:
            fecha = row[columnas["fecha"]]
            if hasattr(fecha, "strftime"):
                fecha = fecha.strftime("%Y-%m-%d")
        
        categoria = None
        if columnas.get("categoria") and columnas["categoria"] in row.index:
            categoria = row[columnas["categoria"]]
        
        concepto_normalizado = concepto.lower().strip()
        concepto_normalizado = re.sub(r'[^\w\s]', ' ', concepto_normalizado)
        concepto_normalizado = ' '.join(concepto_normalizado.split())
        
        movimiento = {
            "id": idx,
            "fecha": str(fecha) if fecha else None,
            "concepto": concepto,
            "importe": round(importe, 2),
            "tipo": tipo,
            "categoria": str(categoria) if categoria else None,
            "conciliado": False,
            "id_extracto_matched": None,
            "concepto_normalizado": concepto_normalizado,
        }
        
        movimientos.append(movimiento)
        
        if tipo == "ingreso":
            total_ingresos += abs(importe)
            num_ingresos += 1
        else:
            total_gastos += abs(importe)
            num_gastos += 1
    
    resumen = {
        "hoja": nombre_hoja,
        "total_movimientos": len(movimientos),
        "num_ingresos": num_ingresos,
        "num_gastos": num_gastos,
        "total_ingresos": round(total_ingresos, 2),
        "total_gastos": round(total_gastos, 2),
        "saldo_neto": round(total_ingresos - total_gastos, 2),
    }
    
    return movimientos, resumen


import io

ProcesarExcelContable = leer_excel_contable
