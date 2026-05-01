from fastapi import FastAPI, UploadFile, File, HTTPException
from fastapi.responses import StreamingResponse, JSONResponse
from fastapi.middleware.cors import CORSMiddleware
import pandas as pd
import io
import json
import re
import sys
import os
from datetime import datetime
from typing import Dict, List, Optional
import tempfile
from openpyxl import load_workbook, Workbook

project_root = os.path.dirname(os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))))
if project_root not in sys.path:
    sys.path.insert(0, project_root)

backend_path = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
if backend_path not in sys.path:
    sys.path.insert(0, backend_path)

from app.procesamiento.clasificador import ClasificadorMovimientos
from app.procesamiento.procesar_excel_contable import leer_excel_contable, detectar_hoja_por_mes
from app.procesamiento.logica_conciliacion import conciliar_movimientos
from app.procesamiento.generar_excel import crear_excel_actualizado, crear_excel_resumen

app = FastAPI()

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

CLASIFICADOR_CACHE = None

def get_clasificador():
    global CLASIFICADOR_CACHE
    if CLASIFICADOR_CACHE is None:
        try:
            from backend.app.procesamiento.clasificador import ClasificadorMovimientos
        except ImportError:
            from app.procesamiento.clasificador import ClasificadorMovimientos
        CLASIFICADOR_CACHE = ClasificadorMovimientos()
    return CLASIFICADOR_CACHE

def detectar_columnas(df: pd.DataFrame) -> Dict[str, str]:
    cols = list(df.columns)
    resultado = {"fecha": None, "concepto": None, "importe": None, "saldo": None}

    for col in cols:
        col_lower = col.lower().strip()
        if resultado["fecha"] is None and ("fecha" in col_lower or "date" in col_lower):
            resultado["fecha"] = col
        if resultado["concepto"] is None and (
            "concepto" in col_lower or "observaciones" in col_lower or "descripcion" in col_lower
        ):
            resultado["concepto"] = col
        if resultado["importe"] is None and col_lower == "importe":
            resultado["importe"] = col
        if resultado["saldo"] is None and col_lower == "saldo":
            resultado["saldo"] = col

    if resultado["importe"] is None:
        for col in cols:
            if pd.api.types.is_numeric_dtype(df[col]):
                col_lower = col.lower()
                if "codigo" not in col_lower and "code" not in col_lower:
                    resultado["importe"] = col
                    break
    return resultado

def limpiar_importe(valor) -> float:
    if pd.isna(valor):
        return 0.0
    if isinstance(valor, (int, float)):
        return float(valor)
    texto = str(valor).strip().replace(".", "").replace(",", ".")
    texto = re.sub(r"[^\d.\-]", "", texto)
    try:
        return float(texto)
    except:
        return 0.0

def normalizar_fecha(fecha) -> str:
    if not fecha:
        return None
    if isinstance(fecha, str):
        return fecha[:10] if len(fecha) >= 10 else fecha
    try:
        return pd.to_datetime(fecha).strftime("%Y-%m-%d")
    except:
        return str(fecha)

_EXTRACTO_DATA = {"movimientos": None, "mes": None, "año": None}
_EXCEL_CONTABLE_DATA = {"movimientos": None, "resumen": None, "contenido": None}

@app.get("/")
def root():
    return {"mensaje": "API de procesamiento de extractos bancarios"}

@app.post("/api/procesar-dos-archivos")
async def procesar_dos_archivos(extracto: UploadFile = File(...), gastos: UploadFile = File(...)):
    contenido_extracto = await extracto.read()
    try:
        if extracto.filename.lower().endswith(".csv"):
            try:
                df_extracto = pd.read_csv(io.StringIO(contenido_extracto.decode("utf-8")))
            except:
                df_extracto = pd.read_csv(io.StringIO(contenido_extracto.decode("latin-1")))
        else:
            df_extracto = pd.read_excel(io.BytesIO(contenido_extracto))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error extracto: {str(e)}")
    
    contenido_gastos = await gastos.read()
    try:
        if gastos.filename.lower().endswith(".csv"):
            try:
                df_gastos = pd.read_csv(io.StringIO(contenido_gastos.decode("utf-8")))
            except:
                df_gastos = pd.read_csv(io.StringIO(contenido_gastos.decode("latin-1")))
        else:
            df_gastos = pd.read_excel(io.BytesIO(contenido_gastos))
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Error gastos: {str(e)}")
    
    df_combined = pd.concat([df_extracto, df_gastos], ignore_index=True)
    columnas = detectar_columnas(df_combined)
    
    if columnas["importe"] is None:
        raise HTTPException(status_code=400, detail="Sin columna de importe")
    
    clasificador = get_clasificador()
    movimientos = []
    for _, row in df_combined.iterrows():
        concepto = str(row.get(columnas["concepto"], "")) if columnas["concepto"] else ""
        importe = limpiar_importe(row.get(columnas["importe"], 0))
        if importe == 0: continue
        
        fecha = row.get(columnas["fecha"]) if columnas["fecha"] else None
        clas = clasificador.clasificar(concepto, importe)
        
        movimientos.append({
            "fecha": str(fecha) if fecha else None,
            "concepto": concepto,
            "importe": round(importe, 2),
            "categoria": clas["categoria"],
            "tipo": clas["tipo"],
            "confianza": clas["confianza"],
        })
    
    total_ingresos = sum(m["importe"] for m in movimientos if m["importe"] > 0)
    total_gastos = sum(abs(m["importe"]) for m in movimientos if m["importe"] < 0)
    
    resumen_cat = {}
    for m in movimientos:
        cat = m["categoria"]
        if cat not in resumen_cat: resumen_cat[cat] = {"ingresos": 0, "gastos": 0}
        if m["importe"] > 0: resumen_cat[cat]["ingresos"] += m["importe"]
        else: resumen_cat[cat]["gastos"] += abs(m["importe"])

    df_csv = pd.DataFrame(movimientos)
    csv_output = io.StringIO()
    df_csv.to_csv(csv_output, index=False, encoding="utf-8")
    
    return {
        "nombre_archivo": extracto.filename,
        "resumen_general": {
            "total_ingresos": round(total_ingresos, 2),
            "total_gastos": round(total_gastos, 2),
            "saldo_neto": round(total_ingresos - total_gastos, 2),
        },
        "movimientos_clasificados": movimientos,
        "resumen_categorias": resumen_cat,
        "csv_contenido": csv_output.getvalue(),
    }

@app.post("/api/upload-extracto")
async def upload_extracto(file: UploadFile = File(...), mes: int = 1, año: int = 2024):
    contenido = await file.read()
    try:
        if file.filename.lower().endswith(".csv"):
            try:
                df = pd.read_csv(io.StringIO(contenido.decode("utf-8")))
            except:
                df = pd.read_csv(io.StringIO(contenido.decode("latin-1")))
        else:
            df = pd.read_excel(io.BytesIO(contenido))
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))
    
    cols = detectar_columnas(df)
    if cols["importe"] is None:
        raise HTTPException(status_code=400, detail="Sin columna de importe")
    
    clasificador = get_clasificador()
    movimientos = []
    for idx, row in df.iterrows():
        concepto = str(row.get(cols["concepto"], "")).strip() if cols["concepto"] else ""
        importe = limpiar_importe(row.get(cols["importe"], 0))
        if importe == 0: continue
        
        clas = clasificador.clasificar(concepto, importe)
        fecha = normalizar_fecha(row.get(cols["fecha"])) if cols["fecha"] else None
        
        movimientos.append({
            "id": idx,
            "fecha": fecha,
            "concepto": concepto,
            "importe": round(importe, 2),
            "tipo": clas["tipo"],
            "categoria": clas["categoria"],
            "confianza": clas["confianza"],
            "concepto_normalizado": concepto.lower(),
        })
    
    _EXTRACTO_DATA.update({"movimientos": movimientos, "mes": mes, "año": año})
    return {"estado": "ok", "movimientos": movimientos}

@app.post("/api/upload-excel-contable")
async def upload_excel_contable(file: UploadFile = File(...), mes: Optional[int] = None, año: Optional[int] = None):
    contenido = await file.read()
    if mes is None:
        mes = _EXTRACTO_DATA["mes"] if _EXTRACTO_DATA["mes"] else 1
    if año is None:
        año = _EXTRACTO_DATA["año"] if _EXTRACTO_DATA["año"] else datetime.now().year

    movs, res = leer_excel_contable(contenido, mes, año)
    _EXCEL_CONTABLE_DATA.update({"movimientos": movs, "resumen": res, "contenido": contenido})
    return {"estado": "ok", "resumen": res, "movimientos": movs}

@app.post("/api/conciliar")
async def conciliar():
    if not _EXTRACTO_DATA["movimientos"] or not _EXCEL_CONTABLE_DATA["movimientos"]:
        raise HTTPException(status_code=400, detail="Faltan datos")
    
    resultado = conciliar_movimientos(_EXTRACTO_DATA["movimientos"], _EXCEL_CONTABLE_DATA["movimientos"])
    return {"estado": "ok", **resultado}

@app.post("/api/generar-excel-actualizado")
async def generar_excel_actualizado_endpoint():
    contenido = _EXCEL_CONTABLE_DATA["contenido"]
    if not contenido: raise HTTPException(status_code=400, detail="No hay Excel")
    
    wb = load_workbook(io.BytesIO(contenido))
    nombre_hoja = wb.sheetnames[0]
    res_conciliacion = conciliar_movimientos(_EXTRACTO_DATA["movimientos"], _EXCEL_CONTABLE_DATA["movimientos"])
    
    excel_bytes = crear_excel_actualizado(
        contenido, nombre_hoja, [], res_conciliacion, 
        _EXTRACTO_DATA["mes"], _EXTRACTO_DATA["año"]
    )
    
    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Actualizado.xlsx"}
    )

@app.post("/generar-informe-final")
async def generar_informe_final(formato: str = "excel"):
    res_conciliacion = conciliar_movimientos(_EXTRACTO_DATA["movimientos"], _EXCEL_CONTABLE_DATA["movimientos"])
    excel_bytes = crear_excel_resumen(_EXTRACTO_DATA["mes"], _EXTRACTO_DATA["año"], res_conciliacion)
    
    return StreamingResponse(
        io.BytesIO(excel_bytes),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=Informe.xlsx"}
    )

if __name__ == "__main__":
    import uvicorn
    uvicorn.run(app, host="127.0.0.1", port=8000)